import openpyxl
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import json




def add_totals(dataframe):

    container = []
    for label, sub_df in dataframe.groupby(['Company']):
        sub_df = sub_df.reset_index(drop=True)
        l = sub_df.shape[0]
        sub_df.loc[l+1, 'Company'] = f'{label[0]} Total'
        sub_df.loc[l+1, ['Ad Value', 'PR Value', 'Count']] = sub_df[['Ad Value', 'PR Value', 'Count']].sum()
        container.append(sub_df)

    dataframe_result = pd.concat(container)
    l = dataframe_result.shape[0]
    dataframe_result.loc[l+1, 'Company'] = 'Grand Total'
    dataframe_result.loc[l+1, ['Ad Value', 'PR Value', 'Count']] = dataframe[['Ad Value', 'PR Value', 'Count']].sum()

    return dataframe_result.fillna('')


def hyperlink_title(REPORT_FILE, sht_name):

    # color of the title with hyperlink
    ft = Font(color="000000FF", underline='single')

    wb = load_workbook(REPORT_FILE)
    ws = wb[sht_name]

    headers = []
    for row in ws.iter_rows(min_row=7, max_row=7):
        for cell in row:
            headers.append(cell.value)
    
    index_title = headers.index('Title')
    index_v3 = headers.index('v3 - Link')

    for row in ws.iter_rows(min_row=8):
        row[index_title].hyperlink = row[index_v3].value
        row[index_title].font = ft

    # remove the link column
    ws.delete_cols(index_v3+1)

    wb.save(REPORT_FILE)
    wb.close()

    return


def del_sheet1(REPORT_FILE):

    wb = openpyxl.load_workbook(REPORT_FILE)
    del wb['Sheet']
    wb.save(REPORT_FILE)
    wb.close()

    return REPORT_FILE




# add sheet header
def add_header(REPORT_FILE, sht_name, company_name, report_month, report_year):
    '''
    REPORT_FILE:=name of xlsx workbook
    sht_name:=name of sheet
    company_name:=name of company
    report_month:=report month
    report_year:=report year 
    '''

    wb = load_workbook(REPORT_FILE)
    ws = wb[sht_name]

    ws['A1'].value = company_name.upper()
    ws['A2'].value = f'Media Meter Inc. Report for {report_month} {report_year}'
    ws['A5'].value = sht_name.title()

    # ws.column_dimensions.group(start='M', end='XFD', hidden=True)

    wb.save(REPORT_FILE)
    wb.close()

    return

@st.cache_resource
def table_formatters():
    
    # set font
    header_font = Font(
        name='Arial',
        size=11,
        bold=True,
        color='00FFFFFF'
    )

    # set fill
    header_fill = PatternFill(
        fill_type='solid',
        start_color='000066CC',
        end_color='000066CC'
    )

    return header_font, header_fill

# format sheet
def format_tableheaders(REPORT_FILE, sht_name):

    wb = load_workbook(REPORT_FILE)
    ws = wb[sht_name]

    # assign cell to variable
    a1 = ws['A1']
    # set font for sheet header
    font = Font(
        name='Arial',
        size=24,
        bold=True,
        color='000066CC'
    )
    # apply font prarameters
    a1.font = font

    # assign cell to variable
    a2 = ws['A2']
    a5 = ws['A5']
    # set font for sheet sub header
    font = Font(
        name='Arial',
        size=11,
        bold=True,
        color='000066CC'
    )
    # apply font prarameters
    a2.font = font
    a5.font = font

    sheet_headers = []
    for row in ws.iter_rows(min_row=7, max_row=7, min_col=1):
        for cell in row:
            sheet_headers.append(cell.value)

    # get table formatters
    header_font, header_fill = table_formatters()  

    for row in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=len(sheet_headers)):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
    
    # set font for table body
    font = Font(
        name='Arial',
        size=11,
        bold=False,
        color='00000000'
    )

    alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    # format DATE column.
    date_column_index = sheet_headers.index('Date')

    for row in ws.iter_rows(min_row=8, min_col=date_column_index+1, max_col=date_column_index+1):
        for cell in row:
            cell.number_format = 'MMM DD, YYY'
            cell.alignment = alignment

    # format length, ave and pr columns
    len_column_index = sheet_headers.index('Length')

    for row in ws.iter_rows(min_row=8, min_col=len_column_index+1, max_col=len_column_index+3):
        for cell in row:
            cell.number_format = '#,##0.00'
    
    # align count column
    for row in ws.iter_rows(min_row=8, min_col=len(sheet_headers), max_col=len(sheet_headers)):
        for cell in row:
            cell.alignment = alignment
    
    # autofit text in cells
    col_number = 1
    for col in ws.iter_cols(min_row=7, min_col=1, max_col=11):
        max_length = 0
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        
        width_to_use = (max_length+1) * 1.2
        col_name = get_column_letter(col_number)
        ws.column_dimensions[col_name].width = width_to_use
        col_number += 1
    
    wb.save(REPORT_FILE)
    wb.close()

    return


def format_datatables(FILE, SHEET, TABLE_LOCATION, TABLE_SIZE):

    wb = load_workbook(FILE)
    ws = wb[SHEET]

    # whole table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    font = Font(
        name='Arial',
        size=11,
        bold=False,
        color='00000000'
    )

    alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    for row in ws.iter_rows(min_row=TABLE_LOCATION[0]+1, min_col=TABLE_LOCATION[1]+1, max_row=TABLE_LOCATION[0]+TABLE_SIZE[0]+1, max_col=TABLE_LOCATION[1]+TABLE_SIZE[1]+1):
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_border

            if type(cell.value)==float:
                cell.number_format = '#,##0.00'
            elif type(cell.value)==int:
                cell.number_format = '#,##0'
    
    if ws.title == 'Share of Voice':
        col_number = 14
        for col in ws.iter_cols(min_row=1, min_col=col_number):
            col_name = get_column_letter(col_number)
            ws.column_dimensions[col_name].width = 15
            for cell in col:
                if type(cell.value) == str:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            col_number += 1
        
        col_number = 9
        for col in ws.iter_cols(min_row=1, min_col=col_number, max_col=col_number+2):
            col_name = get_column_letter(col_number)
            if col_number != 9:
                ws.column_dimensions[col_name].width = 15
            else:
                max_length = 0
                for cell in col:
                    if cell.value != 'Company':
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                
                width_to_use = (max_length+1) * 1.2
                col_name = get_column_letter(col_number)
                ws.column_dimensions[col_name].width = width_to_use
            col_number += 1


    
    # table headers
    font = Font(
        name='Arial',
        size=11,
        bold=True,
        color='00FFFFFF'
    )

    fill = PatternFill(
        fill_type='solid',
        start_color='000066CC',
        end_color='000066CC'
    )

    for row in ws.iter_rows(min_row=TABLE_LOCATION[0]+1, min_col=TABLE_LOCATION[1]+1, max_row=TABLE_LOCATION[0]+1, max_col=TABLE_LOCATION[1]+TABLE_SIZE[1]+1):
        for cell in row:
            cell.fill = fill
            cell.font = font

    wb.save(FILE)
    wb.close()

    return