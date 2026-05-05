import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def sheet_header(ws, header_title):

    thin = Side(border_style="thin", color="000000")

    h1 = ws.cell(row=1, column=1)
    h1.value = 'MCG CORPORATE COMMUNICATIONS'
    h1.font = Font(name='Calibri', size=11, bold=True)
    
    h2 = ws.cell(row=2, column=1)
    h2.value = 'BDO PR REPORT'
    h2.font = Font(name='Calibri', size=11, bold=True)

    header = ws.cell(row=4, column=1)
    header.value = header_title
    header.font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.fill = PatternFill('solid', fgColor='993300')
    header.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    return



def sheet_sub_header(ws, min_row, max_row, min_col, max_col):

    thin = Side(border_style="thin", color="000000")

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cl in row:
            cl.font = Font(bold=True, name='Calibri', size=11, color='000000')

    return


def sheet_body(ws, min_row, min_col, max_col):

    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        
        if row[0].value != None:
            for cl in row:
                cl.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cl.alignment = Alignment(horizontal='center', vertical='center')
        elif row[0].value == 'TOTAL' or row[0].value == None:
            break

    return


def individual_sheet_formatting(xlsx_file, sheet_name, pr_title):

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb[sheet_name]

    # format sheet header
    sheet_header(ws, pr_title)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=17)

    # format sheet sub header
    sheet_sub_header(ws, min_row=5, max_row=5, min_col=1, max_col=17)
    
    # format body
    sheet_body(ws, min_row=6, min_col=1, max_col=17)

    # format total row
    my_row = 1
    for row in ws.iter_rows():
        if row[0].value == 'TOTAL':
            for cl in row:
                cl.font = Font(bold=True, size=14, color='FF0000')
                cl.fill = PatternFill('solid', fgColor='FFFF00')            
            break
        else:
            my_row += 1

    ws.merge_cells(start_row=my_row, start_column=1, end_row=my_row, end_column=6)
    ws.cell(row=my_row, column=1).alignment = Alignment(horizontal='center')
    
    # format tally table
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=6, min_col=1, max_col=2):
        if row[0].value != None:
            for cl in row:
                cl.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # cl.alignment = Alignment(horizontal='center', vertical='center')
    
    # format columns containing values
    for row in ws.iter_rows(min_row=6, min_col=7, max_col=8):
        for cl in row:
            cl.number_format = '₱#,000.00'
    
    # add hyoperlink
    for row in ws.iter_rows(min_row=6, min_col=9, max_col=9):
        for cl in row:
            if cl.value != None:
                my_link = cl.value
                cl.value = 'LINK'
                cl.hyperlink = my_link
                cl.font = Font(color='0000FF', underline='single')
            else:
                break

    wb.save(xlsx_file)
    wb.close()

    return


def sheet_formating(xlsx_file):

    # open workbook
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb['SUMMARY']

    # format sheet header
    sheet_header(ws, 'SUMMARY')
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12)

    # format sheet sub header
    sheet_sub_header(ws, min_row=5, max_row=5, min_col=1, max_col=12)
    
    # format body
    sheet_body(ws, min_row=6, min_col=1, max_col=12)
    
    # change number format
    for row in ws.iter_rows(min_row=6, min_col=8, max_col=9):
        for cl in row:
            cl.number_format = '₱#,000.00'
    
    # change number format
    for row in ws.iter_rows(min_row=6, min_col=11, max_col=12):
        for cl in row:
            cl.number_format = '₱#,000.00'

    # format tally row
    my_row = 5
    for row in ws.iter_rows(min_row=5, min_col=1, max_col=12):
        if row[0].value == 'TOTAL':
            for cl in row:
                cl.font = Font(bold=True, size=14, color='FF0000')
                cl.fill = PatternFill('solid', fgColor='FFFF00')
        else:
            my_row += 1
    
    # merge last row of table containing total
    ws.merge_cells(start_row=my_row, start_column=1, end_row=my_row, end_column=5)

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 90
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20

    wb.save(xlsx_file)


    return