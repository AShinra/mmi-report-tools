import streamlit as st
from datetime import datetime
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
import streamlit_shadcn_ui as ui


def sheet_formatting(wb, header):

    ws = wb['FOODPANDA']
    
    # formatting of header
    ws['A1'].value = header.upper()
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=9)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(name='Nunito', size=18, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='993300')

    # formatting of sub headers
    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font = Font(name='Arial', size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FF6600')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    nsmmyy = NamedStyle(name="cd1", number_format="MM-DD-YYY")
    # formatting of date and media name column
    for row in ws.iter_rows(min_row=4, min_col=1, max_col=9):
        # date column
        row[0].style = nsmmyy
        row[0].font = Font(name='Arial', size=11)
        row[0].fill = PatternFill('solid', fgColor='FFCC99')
        # media name column
        row[1].font = Font(name='Arial', size=11)
        row[1].fill = PatternFill('solid', fgColor='FFCC99')
        # type column
        row[4].font = Font(name='Arial', size=11)
        row[4].fill = PatternFill('solid', fgColor='FFCC99')
        # ave column
        row[7].font = Font(name='Arial', size=11)
        row[7].number_format = '#,000.00'
        # pr column
        row[8].font = Font(name='Arial', size=11)
        row[8].number_format = '#,000.00'

        row[2].font = Font(name='Arial', size=11)
        row[3].font = Font(name='Arial', size=11)
        row[5].font = Font(name='Arial', size=11)
        row[6].font = Font(name='Arial', size=11)

    # add hyperlinks
    for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):
        row[0].hyperlink = row[0].value
        row[0].font = Font(color='0000FF', underline='single')
    
    # set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    return wb


def foodpanda_initial(FILE):

    date_range = ui.date_picker(label='Select Date Range', mode='range', key='foodpanda_date_range', default_value=None)

    if date_range != None:

        st.button(label='Submit', key='submit_foodpanda_data')
        
        if st.session_state['submit_foodpanda_data']:

            foodpanda(date_range, FILE)

    return



def foodpanda(date_range, raw_file):

    # create header and file name for excelfile
    start_date = datetime.strptime(date_range[0], '%Y-%m-%d')
    start_date = start_date.strftime('%b %d')
    end_date = datetime.strptime(date_range[1], '%Y-%m-%d')
    end_date = end_date.strftime('%b %d, %Y')
    
    header = f'FOODPANDA {start_date} to {end_date}'
    f_name = f'FOODPANDA_WEEKLY_REPORT {start_date} to {end_date}'

    # create a temporary excel file
    result_file = Path(__file__).parent/f'temp.xlsx'
    wb = openpyxl.Workbook(result_file)
    wb.save(result_file)
    wb.close()

    # create dataframe
    df = pd.read_excel(raw_file)
    # st.dataframe(df)

    # filter company articles by category
    new_df = df.groupby('Category').get_group('Foodpanda Phils.').reset_index(drop=True)
    # st.dataframe(new_df)

    # get needed columns and re order
    new_df = new_df[['Date', 'Publication', 'v3 - Link', 'Title', 'Media Type', 'Edition', 'Length', 'Ad Value', 'PR Value']]
    # st.dataframe(new_df)

    # rename columns
    new_df.rename(columns={
        'Date':'DATE',
        'Publication':'MEDIA NAME',
        'v3 - Link':'LINK',
        'Title':'TITLE',
        'Media Type':'TYPE',
        'Edition':'EDITION',
        'Length':'LENGTH',
        'Ad Value':'AVE',
        'PR Value':'PR VALUE'}, inplace=True)
    
    # st.dataframe(new_df)

    # save to excel file
    writer = pd.ExcelWriter(result_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    # writer = pd.ExcelWriter(result_file, engine='openpyxl')
    new_df.to_excel(writer, sheet_name='FOODPANDA', index=False, startrow=2)
    writer.close()

    # remove blank sheet
    wb = openpyxl.load_workbook(result_file)
    del wb['Sheet']

    wb = sheet_formatting(wb, header)
    wb.save(result_file)

    _file = open(result_file, 'rb')
    st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
    st.download_button(label='📥 Download Excel File', data= _file, file_name= f'{f_name}.xlsx')

    return