from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable


# create a bar graph
def bar_chart(REPORT_FILE, sht_name, table_dimension, table_loc, chart_loc, chart_size, chart_name, report_month, report_year, fname, _type):

    wb = load_workbook(REPORT_FILE)
    ws = wb[sht_name]

    table_length = table_dimension[0]
    table_width = table_dimension[1]

    start_row = table_loc[0]
    start_col = table_loc[1]

    # get the values
    values = Reference(ws, min_row=start_row+1, min_col=start_col+2, max_row=start_row+1+table_length, max_col=start_col+table_width+1)

    # get categories
    cats = Reference(ws, min_row=start_row+2, min_col=start_col+1, max_row=start_row+1+table_length, max_col=start_col+1)

    # create chart
    _chart = BarChart()
    _chart.type = "col"
    _chart.style = 2
    _chart.grouping = "stacked"
    _chart.overlap = 100
    _chart.gapWidth = 10
    _chart.title = f'{chart_name} ({_type}) {report_month} {report_year}\n{fname.title()}'
    # place legend at the bottom of the chart
    _chart.legend.position = 'b' 

    # add the values
    _chart.add_data(values, titles_from_data=True)
    _chart.set_categories(cats)
    _chart.shape = 4
    _chart.DataTable = DataTable(True, True, True, True) 
    # place position
    ws.add_chart(_chart, anchor=chart_loc)

    _chart.height = chart_size[0]
    _chart.width = chart_size[1]

    wb.save(REPORT_FILE)
    wb.close()

    return


# create a pie chart
def pie_chart(REPORT_FILE, sht_name, _data, _labels, cht_loc, chart_size, chart_name, report_month, report_year, fname, _type):

    wb = load_workbook(REPORT_FILE)
    ws = wb[sht_name]

    pie = PieChart()

    data = Reference(ws, min_row=_data[0], min_col=_data[1], max_row=_data[2], max_col=_data[3])
    labels = Reference(ws, min_row=_labels[0], min_col=_labels[1], max_row=_labels[2], max_col=_labels[3])

    pie.add_data(data,titles_from_data=True)
    pie.set_categories(labels)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showLeaderLines = True
    pie.title = f'{fname.title()} {chart_name} ({_type})\n{report_month} {report_year}'
                
    pie.height = chart_size[0]
    pie.width = chart_size[1]
    pie.legend = None

    ws.add_chart(pie, cht_loc)

    wb.save(REPORT_FILE)
    wb.close()

    return
    
