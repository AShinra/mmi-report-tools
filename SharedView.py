import streamlit as st
import pandas as pd
import openpyxl
import json
import os
from io import BytesIO
from streamlit_option_menu import option_menu
from pathlib import Path
from Tools import del_sheet1
from Sheets.Daily_Statistics import daily_stats
from Sheets.Single_Sheet import single_sheet_report
from Sheets.Multiple_Sheet import multiple_sheet_report
from Sheets.Media_Breakdown import media_stats
from Sheets.SOV import my_sov
from Charts.charts import bar_chart, pie_chart
import streamlit_shadcn_ui as ui
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

_columns = {
    'Date':12,
    'Title':70,
    'Influencer':25,
    'Engagement Score':22
}

# font formats
sheet_header_font = Font(name='Arial', color="000000FF", size=25, bold=True)
sheet_subheader_font = Font(name='Arial', size=14, bold=True)
table_header_font = Font(name='Arial', size=12, bold=True)
table_body_font = Font(name='Arial', size=10)
hyperlink_font = Font(name='Arial', size=10, underline='single', color="000000FF", bold=True)

# table header fills
header_fills = {
    'Facebook':PatternFill(start_color='000000FF',end_color='000000FF',fill_type='solid'),
    'Instagram':PatternFill(start_color='00FF99CC',end_color='00FF99CC',fill_type='solid'),
    'Twitter':PatternFill(start_color='00000000',end_color='00000000',fill_type='solid'),
    'Youtube':PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
}

header_fontcolors = {
    'Facebook':Font(color='00FFFFFF',name='Arial', size=12, bold=True),
    'Instagram':Font(color='00FFFFFF',name='Arial', size=12, bold=True),
    'Twitter':Font(color='00FFFFFF',name='Arial', size=12, bold=True),
    'Youtube':Font(color='00FFFFFF',name='Arial', size=12, bold=True)
}

platform_imgs = {
    'Facebook':'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAADFUlEQVR4Ab2XA4ycURSFJ52pbdu27TZOEau2bRtRGaO2g9q2bdva3a8nufXg/cubfLuDO++eh3vmjc9LpOpKQJQSvcRqcUW8E1EG78RVsUb0FqVFwDGs58I1xAJx24qBgyjLZaGoKQJxLZ5bTBfPBHHkmZgh8sS2eDWxV8QIEoB9oobX4o3EFYEXUnYxUnSG5AYphL3+dy5XRRNX8epW3FvhgBUie18oPhwqjIFyo6HIUHstTbeQImqEK57Hlh1cpOhigzebBQt2wdFbcOcFPHkLj9/AzWdw/Ba0nW8rEmI78v5fPJmY7nXmWXrDtC3w+iMRY9AK8HcKOc5M4f9bQE3x1F3c/o9eB9+iAOIs4Jmo9au43/ocImPLWWkc3HsJEA8BxiLh95lrcUvgItDJBg0VMTHw6DXsugzrT8Km09Au9BkwzKzK+sxe+e5l+VPr/9LDhIwdF6HaBMjY0w5o2u5BXRDKMfv4zNu9Hb5MvWDPFYLiwxdoPReSdbS8v3GMu87nve+tt4/dIihuP4diw8yALN8z1ySAt67C5nKQrQ+cuE1QXH0M+QdBoJPlGp5W4L0EEBWpeIYeUHSYzbDyODh3n6C49QxqT7YcOaKwz+izLhExEQTYrGtOgksP4e4La78v3wkKeQL3X5kbCuXByTtQZpSN4RLwNlLbNZgG7z4T6zh7D/IMMNt2bcGVxBCw/pS1omMLrksAaxJDwIxttvwS4GzD3mGMSAOYuRy+YYfvwgP4+JWg+PLNzsnZe5Z3VrSfbxNwXN36SgBlxK3QXWDLmKu/UWIEnLlHUNx4ChXGKqcf5O5vpHd3wJ1fVuwXCyP6gHmBjCi8DxQYBFoxy/XmgouF3+PXsQ2YI5yAJ1BwsBX2flml9t/3Ab+YkYQCZgl/qCvZviQQsF/kDXcvrCGuJqKAa6Km62bcWFxNcAFWvGlwxfArsT8BBRxwzDzsNX2meB4PAc/FLJE3Pj9Oa4pFGvyOBER5EBAl7lifUyvBfiXLVsvkHUifiw9ZB1wX70UM4voT3hcewnUZ0Trl9hVlvBb+AQNcQrB9dXA4AAAAAElFTkSuQmCC',
    'Instagram':Font(color='00FFFFFF',name='Arial', size=12, bold=True),
    'Twitter':'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAKS0lEQVR4AZ2UA5RkS7ZA94mIq8ys6mrb7v7PGNu2bdu2bdvssZ9trme03V1VWZWZFxFxfv5a41n6s8/al2EJ/Ccve8P566a78YSF6+Ys745PLW2bbA42LSQQbFQVNWJjUKIniRZRkOiNiSImqiRirdZ+YLNkl4HbxvdOTnSS7CbgGv4N+ehTLuFvjK+IW6/faT/nPfeAlEG3ouUcje9jYsAhmCgkUTExIgquAadgQ8QEcDFim4ioINFgsbRahlgGMPGizfdY9Lx/boh892XXAXCl7HzktVeNbUvzhCJTUtNgQwQFh5JKQNQgPpLEgFOQCDYG0iCYAAaP/duIeMBHiIpEwUQIA5jqKusfMu/RwDYA5yf7zH90cuyN781/PtI2jJg+Ogg4DSQGRJXqkKc3WSJ1iYlg1cw0wkbFhIBvFIvBxohDEAxZO6G9IEfEEGsQY7GFkAw8+3974OcP/uwxJwOXyWl/uCn/9Zf3f/32K6onLVlt0SZgRUmc4eitk8TxktX/M8qCFW06bcElCUbMTMUOj0TFqMEOVQ/BN9TdQG97j0PXjlNkKWOrRmh8RFWwScLE7QNGThz95aanLHqm23fN3gf0dvQetXA0IeuVGCAxhv0X72PtsW0e99W7svm+K/hvuO203Zz72gvoXrmPuWvm4xUolblzDH7X4CGDm/v3czecOXVHM9C8VUSSKpJnjn0X7+Buj1rGc37yYAAgcPC6w/T2l2it4BUNEdUIgBGLNYKkik0c2dI28zbOZc19lrHmisfxu8f9il0/3c2c1YsI2oATTO2d3xmOcWHCLO/EmrSJpAambhxn6yljw8ofBMBl37mCC999KdWeAdIEbIyYqFgEiFgiZiaY0TrLtJ+goMOdv/FQNjxzCw/6yUP55ak/oHfJEVorRolRQRWO9DuOXjdr2QTnA84Gku44937ZKYBw3U+v4cdP/wlLs7nMX9TG/q1Ca3ABpO8BJckcKgbjoTrSZ+2jTqTf7/OHZ30Vlz+fNU/cxJZXHsuFT/kTrilQI2hUqJrcdZw2sazIBOrD0yxb3+aYJ20C4OJPnMuqZJQFK1uEqiRxFlMr5a79VLEkT9uICr2mpDAZTTQseuQW7vTzRwIgD6y47IU/Hzbgzax98mZufsNF6KEedl5OJGLKsnCZjbVvarJUaCammbd+HohhfPshwnVHmTs/w1Q9cmvx3S6DQ/tYfq8trHzBnRk5dhFGDUev2c/Or1zIntPOJx5dxN9ouZSJyUl6uyZoLx+jWJHT232IRHMQkEHpXBJrmTlMiORNn86oAFDvmyQru2QjYzgnMDVNPHSAO33pCax8/j35Z0a2LGDlE47llq8ew1XP+wbnn/ppGG3TO2M3Hdei2tWdaUA6x9GnxEpFULBk0RUmaDQlmUSiluS5AEAIZFRktsGqoXdwByd+7okzlQcN3Pbh3zLx6+sRNcx+2DGsfsP9WffcO2Gk4prnfolRVtGZPUZvfAppIgBpKlgaTPQYFYxRnKNBhuYEIn0y6wGQGEkZvict/I7DrHjAZla++AEAXPGgD3DkjxdTsAgQDp5/PhNnXMdJf3w1a55zTw7/+EqqP+1EbENCjcQAgBWPxWOQmXuiDSbVmlRKclvTTnokpgRAYiA3FWk6IKkPsfRxJwBw2yd/SvePp7FwzQY6a0fprB5l6bL1HP3TWWz/2O8BWPKYU/B0MRpQPBoDAKIRS42RCoOCCqZIBhRuQD40G5raGgBBaaU9cp2imG/pnLwGgMGZFzO7NYrJexgzjWOAcQ2jtOmddiUA7TsuJ0lybF1hqCAoAFbAILgQMTSoqXFJUmKykiz12KJPkg0AcK4mb0+RFQWmPQAbAXC2gqzEpCUIIA3ESGoHWAYAiAacGb7TwhEQUQCiRoSAAVwMmOjFZEmlrbSkyIe2+qRZBYDYSNoevo9WuGYf1XXXATB6r+PR8ZvJWgFXeGyrweUVMewY/tsKQP+i25CqiysEQwPmbyPQYKhBAkbBKZhWVtFKB2T50GKKNK0BkBTSTp+kKMnnw9QffwnAwpc+i/aT701z8R/h4PVDb6Z/7WmM3e8uLHndEwA4+tMzyQC1FZaAEQHAEf/6HlCJKKjLskpdMY0rFNupsGkJgE2EZNSTtAZkmxZRXX4a4z/6NrOf8HRWfPuLTNzlJMozLiB6YezOz2H+K54NwK7PbyP8+RyKNZvwZYUxATEAEBUEjxARiagEXDLTgAFpITDaYG2Xv41AMlLjWh4c5BsWM/WZN2ELx+jDn8zcF78Ihv4zB765jalXfIjOyqWEVDF1RG2DJgaAWFUIAVGPiGBiFDcccrGFx+UR02mwOgmAzJ6NG1Nc0UddgnRaJK0l9D/xcspzfkv7QU8lXb0OwVLefBPj3/0p1a/OZmTLWnxeoN0S1JAUgls0C4A4MY3BAQGDRQji0lk+Gh1giwQ3p4XWB1AgW7iWZPkCzMHtyLx10NRI0cHO20q98zy6H/oLks9Fa0PYNwlmhNadNxBLh+kNkJEO5d49ZGtX0Fq7kAaIu46QpCmqEQmKyRNvJPdl0VayvCFZ1Eamb2Fw2S8xQPG41+D9Lky5AzOimKLBtAPpptWkJ2wkWTmbdPkYxYnrybYsQ/IKSRvIa8LeG6m6tzL/DU8BGJ6OZw0bvp1kLEeiMhNOS2dmtyuJ+3GZhdyQaoG/4MvoSY9g5PhHYF7xU8JZn0F7ByAGCAGpA6YUtInERpCBQB/ilCP2UqgK7JIlLHv/25j7hHvRAAc+8UtyCsSB8QExlqTtKhfNYMJkBjsimBTMnA1w8Bp6215K61GfpX38Y4j/5+GboXeI2FSY4JGgRB/RoBANUQVCgpJgZi8gXb8BEYjAbY9/J3rh1aRrjiNUHoxDiUQ1U7L3L194ZHbVth+mWmbMakPqsNYQjt4AS48hvdNrSRbfAeH/hwLTZ1/A/jd/gfqC20g3HUPVd0Sf4AcJTXRh7DEnPEV2XXNOkl79na+OHLny6TpvBSYJqLOYNIPpXcRQI3M3Q3sRalsIDhOE6APa+KEKZYAKfKnEbo0emqJ/2z4GF92OjbOIi9fgJxJ8PwEpmN4xgTlu/e9XvOCBz5D9X7ySsOHCjbMObLsycXmunTHUKeIcuASLEMtx1E+hIYKfkVh7pFRi3xMHDaHXQFfxk4HQBa1nEdwCwnRCM2Hw/QzftAlHIoPJmlXvecEdgYscwxd7yYk3Du42/Zjk6AW/NXlAkwKxBdFCNALFUmQmgBARHyB4qCK2CpiywfRqdNpjpiN+WglHFR0XmiqimUOqiJ0c0J/wdJ5x78cfvvm6iwDsM1fdhf7RQ/SuaN3cvv/yHzrbbCLL1pgsw1mBNAFrcFYxBowoGMVqRPCICUio/2oDTY0MamK/JDQRug1EhzYpfZ9dyF23PNT3B6cPZShy6SM/xr+z9eNz15h49BjaYwu08R2QZUI6V4lNDF5M8GgIaipvY+2T0NRo6SX2gml6lQ3TTWbLfNztnTw0befvqa+fODDIxq4Cruff+F/YMSYOvVvlzAAAAABJRU5ErkJggg==',
    'Youtube':Font(color='00FFFFFF',name='Arial', size=12, bold=True)
}

def format_compilation(ws, k):

    ws['A1'].font = sheet_header_font
    ws['A2'].font = sheet_subheader_font

    for row in ws.iter_rows(min_row=8, min_col=1):
        row[1].hyperlink = row[2].value

    for row in ws.iter_rows(min_row=7, max_row=7, min_col=1):
        for cl in row:
            cl.font = header_fontcolors[k]
            cl.fill = header_fills[k]
            if cl.value == 'Raw Date':
                cl.value = 'Date'
    
    for row in ws.iter_rows(min_row=8, min_col=1):
        for cl in row:
            cl.font = table_header_font
    
    for row in ws.iter_rows(min_row=8, min_col=1):
        if row[0].value == 'Total':
            row[0].font = table_header_font
            row[-1].font = table_header_font
    
    for row in ws.iter_rows(min_row=8, min_col=2, max_col=2):
        for cl in row:
            cl.font = hyperlink_font
    
    ws.delete_cols(3, 1)
    

    return


def tally_sheet(REPORT_FILE, platform_count, company_name, my_range):

    wb = openpyxl.load_workbook(REPORT_FILE)
    ws = wb.active
    ws.title = 'TALLY'

    # get data from platform_count
    _total = 0
    for k, v in platform_count.items():
        _total += v

    ws['A2'] = 'CLIENT'
    ws['B2'] = 'SM Hotels and Conventions Corporation'
    ws['A3'] = 'ACCOUNT'
    ws['B3'] = company_name
    ws['A4'] = 'RANGE'
    ws['B4'] = my_range
    ws['A5'] = 'TOTAL'
    ws['B5'] = _total

    ws['B7'] = 'FACEBOOK'
    ws['C7'] = 'INSTAGRAM'
    ws['D7'] = 'TWITTER'
    ws['E7'] = 'YOUTUBE'

    ws['A8'] = 'COUNT'
    ws['B8'] = platform_count['Facebook']
    ws['C8'] = platform_count['Instagram']
    ws['D8'] = platform_count['Twitter']
    ws['E8'] = platform_count['Youtube']

    wb.save(REPORT_FILE)
    return


def create_sv_compilation(sv_sheets, REPORT_FILE, company_name, my_range):

    for k, v in sv_sheets.items():
        writer = pd.ExcelWriter(REPORT_FILE, engine='openpyxl', mode='a')
        v.to_excel(writer, sheet_name=k, index=False, startrow=6)
        writer.close()

        wb = openpyxl.load_workbook(REPORT_FILE)
        ws = wb[k]
        ws['A1'] = company_name
        ws['A2'] = f'{k} Compilation for {my_range}'

        

        format_compilation(ws, k)

        wb.save(REPORT_FILE)
        wb.close()

    return


def sv_pivot(dataframe, company_name, platform_list):

    sv_sheets = {}

    index_dict={
        'Facebook':['Raw Date', 'Title', 'Article Source', 'Influencer', 'Engagement Score', 'FB - Likes', 'FB -  Shares', 'FB - Comments ', 'FB -  Wow', 'FB -  Love', 'FB - Sad', 'FB - Angry', 'FB - Haha', 'FB -  Care'],
        'Instagram':['Raw Date', 'Title', 'Article Source', 'Influencer', 'Engagement Score', 'Instagram - Likes ', 'Instagram -  Comments'],
        'Twitter':['Raw Date', 'Title', 'Article Source', 'Influencer', 'Engagement Score', 'Twitter - Retweets ', 'Twitter -  Comments', 'Twitter -  Likes', 'Twitter - Replies'],
        'Youtube':['Raw Date', 'Title', 'Article Source', 'Influencer', 'Engagement Score', 'Youtube -  Likes', 'Youtube - Dislikes', 'Youtube - Comments']
    }

    platform_count = {}
    for platform in platform_list:
        _index = index_dict[platform]
        pvt = dataframe.query('Bucket==@company_name & Platform==@platform').pivot_table(index=_index, aggfunc={'Article ID':'count'}).reset_index()
        pvt.rename(columns={'Article ID': 'Count'}, inplace=True)

        l = pvt.shape[0]

        pvt.loc[l+1, 'Raw Date'] = 'Total'
        pvt.loc[l+1, 'Count'] = pvt['Count'].sum()

        platform_count[platform] = l

        sv_sheets[platform] = pvt

    return sv_sheets, platform_count


def get_plaformt_list(df, company_name):

    platform_list = []
    for i in df.index:
        if df['Bucket'][i] == company_name:
            platform_list.append(df['Platform'][i])
    
    platform_list = list(dict.fromkeys(platform_list))

    return platform_list


def get_company_list(df):

    company_list = []
    for i in df.index:
        if df['Bucket'][i] == 'Industry':
            continue
        else:
            if df['Bucket'][i] not in company_list:
                company_list.append(df['Bucket'][i])
          
    return company_list


def main_sv(cleaned_file, name):

    # get company list from dataframe
    df = pd.read_excel(cleaned_file)
    company_list = get_company_list(df)

    with st.container(border=True):
        st.subheader(':blue[REPORT INFO]')

        col1, col2 = st.columns(2)

        with col1:        
            my_range = ''
            date_range = ui.date_picker(label='', mode='range', key='sv_date_range', default_value=None)
            
            # st.write(st.session_state['sv_date_range'])

            if date_range:
                if date_range[0] == date_range[1]:
                    my_range = datetime.strptime(date_range[0], '%Y-%m-%d').date()
                    my_range = my_range.strftime('%B %d, %Y')
                else:
                    my_range1 = datetime.strptime(date_range[0], '%Y-%m-%d').date()
                    my_range1 = my_range1.strftime('%B %d, %Y')

                    my_range2 = datetime.strptime(date_range[1], '%Y-%m-%d').date()
                    my_range2 = my_range2.strftime('%B %d, %Y')

                    my_range = f'{my_range1} to {my_range2}'
            
            


            # cb_daterange = st.checkbox(label=':blue[Date Range]:red[*]', value=False)

            # if cb_daterange:
            #     report_month = ui.date_picker(label='', mode='range', key='sv_date_range', default_value=None)
            #     st.write(report_month)
            #     report_year = ''

            #     if report_month not in [None, '']:
            #         d_range1 = report_month[0]
            #         d_range2 = report_month[1]

            #         report_year = d_range2.split('-')[0]

            #         start_date = datetime.strptime(d_range1, '%Y-%m-%d')
            #         start_date = start_date.strftime('%b %d')
            #         end_date = datetime.strptime(d_range2, '%Y-%m-%d')
            #         end_date = end_date.strftime('%b %d')

            #         report_month = f'{start_date} to {end_date}'

            # else:
            #     report_month = st.selectbox(':blue[Report Month]:red[*]', options=
            #                                 [
            #                                     'January',
            #                                     'February',
            #                                     'March',
            #                                     'April',
            #                                     'May',
            #                                     'June',
            #                                     'July',
            #                                     'August',
            #                                     'September',
            #                                     'October',
            #                                     'November',
            #                                     'December'
            #                                 ])
                
            #     report_year = st.number_input(label=':blue[Report Year]:red[*]', min_value=2000, value=2024)

        with col2:
            # fname = st.text_input(label=':blue[Client Name]:red[*]')
            company_name = st.selectbox(label=':blue[Client Name]:red[*]', options=company_list)

        # get plaform list
        platform_list = get_plaformt_list(df, company_name)

        st.button('PROCESS', key='sv_submit', use_container_width=True)

        if st.session_state['sv_submit']:
            # create temporary filename used by user
            REPORT_FILE = Path(__file__).parent/f'Temp_Files/sv.xlsx'
            wb = openpyxl.Workbook(REPORT_FILE)
            wb.save(REPORT_FILE)
            wb.close()

            sv_sheets, platform_count = sv_pivot(df, company_name, platform_list)

            # create tally sheet
            tally_sheet(REPORT_FILE, platform_count, company_name, my_range)

            # place platform compilation to sheets
            create_sv_compilation(sv_sheets, REPORT_FILE, company_name, my_range)

            result_file = open(REPORT_FILE, 'rb')
            st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
            st.download_button(label='📥 Download Excel File', data= result_file, file_name=f'SharedView Report - {company_name}.xlsx')
        

        
    
    # with st.container(border=True):
    #     st.subheader(':blue[CATEGORY SELECT]')
                
    #     col1, col2, col3 = st.columns(3)
    #     with col1:
    #         with st.container(border=True):
    #             main_cat = st.multiselect(':orange[Main Category]', options=get_category_list(df), key='main_cat')

    #     with col2:
    #         with st.container(border=True):
    #             comp_cat = st.multiselect(':orange[Competitor Category]', options=get_category_list(df), key='comp_cat')
        
    #     with col3:
    #         with st.container(border=True):
    #             ind_cat = st.multiselect(':orange[Industry Category]', options=get_category_list(df), key='ind_cat')

    # with st.container(border=True):

    #     col1, col2 = st.columns(2)

    #     with col1:
    #         st.subheader(':blue[SHEET LAYOUT]')
    #         main_page = st.checkbox(label='Main Multiple Sheets', key='main_page', value=False)
    #         comp_page = st.checkbox(label='Competitor Multiple Sheets', key='comp_page', value=False)
        
    #     with col2:
    #         st.subheader(':blue[CHARTS]')
    #         dailystats = st.checkbox(label='Daily Statistics', value=True)
    #         mediastats = st.checkbox(label='Media Statistics', value=True)
    #         sov = st.checkbox(label='Share of Voice', value=True)

    # col1, col2, col3 = st.columns(3)
    # with col2:
    #     button1 = st.button('📊 :blue[Process Report]', use_container_width=True)
    
    # if button1:
    #     # create temporary filename used by user
    #     REPORT_FILE = Path(__file__).parent/f'Temp_Files/{name}.xlsx'
    #     wb = openpyxl.Workbook(REPORT_FILE)
    #     wb.save(REPORT_FILE)
    #     wb.close()

    #     if main_cat != []:
    #         if main_page:
    #             multiple_sheet_report(df, main_cat, company_list, fname, report_month, report_year, REPORT_FILE)
    #         else: 
    #             single_sheet_report(df, main_cat, 'main', fname, report_month, report_year, REPORT_FILE)
        
    #     if comp_cat != []:
    #         if comp_page:
    #             multiple_sheet_report(df, comp_cat, company_list, fname, report_month, report_year, REPORT_FILE)
    #         else:
    #             single_sheet_report(df, comp_cat, 'comp', fname, report_month, report_year, REPORT_FILE)
    #     else:
    #         st.warning(f':red[NOTE:] No Competitor Sheet')
        
    #     if ind_cat != []:
    #         single_sheet_report(df, ind_cat, 'ind', fname, report_month, report_year, REPORT_FILE)
    #     else:
    #         st.warning(f':red[NOTE:] No Industry Sheet')
        
    #     # remove first sheet with no data
    #     REPORT_FILE = del_sheet1(REPORT_FILE)

    #     # create daily statistics sheet
    #     if dailystats:
    #         # create count and values tables
    #         table1_loc, table1_dimension, table2_loc, table2_dimension = daily_stats(df, main_cat, REPORT_FILE, "Daily Statistics")
    #         # create bar graph
    #         chart_size = [15, 25]
    #         bar_chart(REPORT_FILE, 'Daily Statistics', table1_dimension, table1_loc, 'A1', chart_size, 'Daily Statistics', report_month, report_year, fname, 'Count')
    #         bar_chart(REPORT_FILE, 'Daily Statistics', table2_dimension, table2_loc, 'A35', chart_size, 'Daily Statistics', report_month, report_year, fname, 'Value')
        
    #     # create monthly statistics sheet
    #     if mediastats:
    #         # create count and values tables with pie charts
    #         media_stats(df, main_cat, REPORT_FILE, "Media Statistics", report_month, report_year, fname)

    #     if sov:
    #         sov_catlist = main_cat + comp_cat
    #         my_sov(df, sov_catlist, REPORT_FILE, report_month, report_year, fname)    

    #     wb = openpyxl.load_workbook(REPORT_FILE)

    #     for ws in wb.worksheets:
    #         ws.sheet_view.showGridLines=False
        
    #     wb = sheet_formatting(wb)
        
    #     wb.save(REPORT_FILE)
    #     wb.close()


    #     result_file = open(REPORT_FILE, 'rb')
    #     st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
    #     st.download_button(label='📥 Download Excel File', data= result_file, file_name= f'{fname.title()}_Monthly_Statistics_{report_month}_{report_year}.xlsx')


    return   

    
    
    
    
    