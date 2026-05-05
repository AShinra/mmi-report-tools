import streamlit as st
import pandas as pd
import openpyxl
import json
import os
from io import BytesIO
from streamlit_option_menu import option_menu
from pathlib import Path
from Tools import del_sheet1
from Sheets.Daily_Statistics import daily_stats
from Sheets.Single_Sheet import single_sheet_report
from Sheets.Multiple_Sheet import multiple_sheet_report
from Sheets.Media_Breakdown import media_stats
from Sheets.SOV import my_sov
from Sheets.WordCloud import wc
from Charts.charts import bar_chart, pie_chart
import streamlit_shadcn_ui as ui
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from Common import read_excel_file


def sheet_formatting(wb):

    # format tables in Daily Stats Sheet
    try:
        ws = wb['Daily Statistics']

        for rw in ws.iter_rows(min_row=36, min_col=18):
            for cl in rw:
                cl.number_format = '#,000.00'
                cl.alignment = Alignment(horizontal='right')
        
        for cl in ws.iter_cols(min_row=36, min_col=17):
            cl_name = get_column_letter(cl[0].column)
            ws.column_dimensions[cl_name].width = 15
    except:
        pass

    try:
        ws = wb['Media Statistics']

        for rw in ws.iter_rows(min_row=2, min_col=13):
            for cl in rw:
                if cl.value == 'Value':
                    continue
                else:
                    cl.number_format = '#,000.00'
                    cl.alignment = Alignment(horizontal='right')
        
        for rw in ws.iter_rows(min_row=27, min_col=11, max_col=11):
            for cl in rw:
                if cl.value in ['Media Type', 'Website', 'Publication', 'Channel', 'Station']:
                    continue
                else:
                    cl_name = get_column_letter(cl.column)
                    cl.alignment = Alignment(horizontal='left')
                    ws.column_dimensions[cl_name].width = 30

        for cl in ws.iter_cols(max_row=1, min_col=12, max_col=13):
            cl_name = get_column_letter(cl[0].column)
            ws.column_dimensions[cl_name].width = 15

    except:
        pass

    return wb
      


def get_category_list(df):

    cat_list = df['Category'].to_list()
    cat_list = list(dict.fromkeys(cat_list))
    cat_list.sort()
    
    return cat_list


def get_company_list(df):

    company_list = []
    for i in df.index:
        if df['Company'][i] not in company_list:
            company_list.append(df['Company'][i])
    
    return company_list


def main(cleaned_file, name):

    # get company list from dataframe
    df = read_excel_file(cleaned_file)
    company_list = get_company_list(df)

    with st.container(border=True):
        st.subheader(':blue[REPORT INFO]')

        col1, col2 = st.columns(2)

        with col1:        
            cb_daterange = st.checkbox(label=':blue[Date Range]:red[*]', value=True)

            if cb_daterange:

                report_month = ui.date_picker(label='Select Date Range', mode='range', key='br_date_range', default_value=None)

                report_year = ''

                if report_month not in [None, '']:
                    d_range1 = report_month[0]
                    d_range2 = report_month[1]

                    report_year = d_range2.split('-')[0]

                    start_date = datetime.strptime(d_range1, '%Y-%m-%d')
                    start_date = start_date.strftime('%b %d')
                    end_date = datetime.strptime(d_range2, '%Y-%m-%d')
                    end_date = end_date.strftime('%b %d')

                    report_month = f'{start_date} to {end_date}'


            else:

                report_month = st.selectbox(':blue[Report Month]:red[*]', options=
                                            [
                                                'January',
                                                'February',
                                                'March',
                                                'April',
                                                'May',
                                                'June',
                                                'July',
                                                'August',
                                                'September',
                                                'October',
                                                'November',
                                                'December'
                                            ])
                
                report_year = st.number_input(label=':blue[Report Year]:red[*]', min_value=2000, value=2024)

        with col2:
            fname = st.text_input(label=':blue[Client/File Name]:red[*]', key='_filename')
    
    with st.container(border=True):
        st.subheader(':blue[CATEGORY SELECT]')
                
        col1, col2, col3 = st.columns(3)
        with col1:
            with st.container(border=True):
                main_cat = st.multiselect(':orange[Main Category]', options=get_category_list(df), key='main_cat')

        with col2:
            with st.container(border=True):
                comp_cat = st.multiselect(':orange[Competitor Category]', options=get_category_list(df), key='comp_cat')
        
        with col3:
            with st.container(border=True):
                ind_cat = st.multiselect(':orange[Industry Category]', options=get_category_list(df), key='ind_cat')


    with st.container(border=True):

        col1, col2 = st.columns(2)

        with col1:
            st.subheader(':blue[SHEET LAYOUT]')
            main_page = st.checkbox(label='Main Multiple Sheets', key='main_page', value=False)
            comp_page = st.checkbox(label='Competitor Multiple Sheets', key='comp_page', value=False)
        
        with col2:
            st.subheader(':blue[CHARTS]')

            cola, colb = st.columns(2)
            with cola:
                dailystats = st.checkbox(label='Daily Statistics', value=True)
                mediastats = st.checkbox(label='Media Statistics', value=True)
            with colb:
                sov = st.checkbox(label='Share of Voice', value=True)
                word_cloud = st.checkbox(label='Word Cloud', value=True)

    col1, col2, col3 = st.columns(3)
    with col2:
        button1 = st.button('📊 :blue[Process Report]', use_container_width=True)
    
    if button1:
        # create temporary filename used by user
        REPORT_FILE = Path(__file__).parent/f'Temp_Files/{name}.xlsx'
        wb = openpyxl.Workbook(REPORT_FILE)
        wb.save(REPORT_FILE)
        wb.close()

        if main_cat != []:
            if main_page:
                multiple_sheet_report(df, main_cat, company_list, fname, report_month, report_year, REPORT_FILE)
            else: 
                single_sheet_report(df, main_cat, 'main', fname, report_month, report_year, REPORT_FILE)
        
        if comp_cat != []:
            if comp_page:
                multiple_sheet_report(df, comp_cat, company_list, fname, report_month, report_year, REPORT_FILE)
            else:
                single_sheet_report(df, comp_cat, 'comp', fname, report_month, report_year, REPORT_FILE)
        else:
            st.warning(f':red[NOTE:] No Competitor Sheet')
        
        if ind_cat != []:
            single_sheet_report(df, ind_cat, 'ind', fname, report_month, report_year, REPORT_FILE)
        else:
            st.warning(f':red[NOTE:] No Industry Sheet')
        
        # remove first sheet with no data
        REPORT_FILE = del_sheet1(REPORT_FILE)

        # create daily statistics sheet
        if dailystats:
            # create count and values tables
            table1_loc, table1_dimension, table2_loc, table2_dimension = daily_stats(df, main_cat, REPORT_FILE, "Daily Statistics")
            # create bar graph
            chart_size = [15, 25]
            bar_chart(REPORT_FILE, 'Daily Statistics', table1_dimension, table1_loc, 'A1', chart_size, 'Daily Statistics', report_month, report_year, fname, 'Count')
            bar_chart(REPORT_FILE, 'Daily Statistics', table2_dimension, table2_loc, 'A35', chart_size, 'Daily Statistics', report_month, report_year, fname, 'Value')
        
        # create monthly statistics sheet
        if mediastats:
            # create count and values tables with pie charts
            media_stats(df, main_cat, REPORT_FILE, "Media Statistics", report_month, report_year, fname)

        if sov:
            sov_catlist = main_cat + comp_cat
            my_sov(df, sov_catlist, REPORT_FILE, report_month, report_year, fname)

        if word_cloud:
            wc(df, REPORT_FILE, main_cat, fname)

        wb = openpyxl.load_workbook(REPORT_FILE)

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines=False
        
        wb = sheet_formatting(wb)
        
        wb.save(REPORT_FILE)
        wb.close()


        result_file = open(REPORT_FILE, 'rb')
        st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
        st.download_button(label='📥 Download Excel File', data= result_file, file_name= f'{fname.title()}_Monthly_Statistics_{report_month}_{report_year}.xlsx')


    return

    

# main script
def basic_report(name):        
    st.header(f'📊 Basic Report')

    cleaned_file = st.file_uploader(label=':blue[Upload Cleaned File]:red[*]', type=["xls","xlsx"], key="Demo", accept_multiple_files=False)

    if cleaned_file:
        main(cleaned_file, name)
    else:
        st.stop()
    
    
    
    
    